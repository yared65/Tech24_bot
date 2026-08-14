import os
import logging
import asyncio
import threading
import json
import urllib.parse
from datetime import datetime, timedelta, timezone
from io import BytesIO
from flask import Flask

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes

# ==========================================
# 1. LOGGING & CONFIGURATION
# ==========================================
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
EMAIL = os.environ.get("EMAIL")
PASSWORD = os.environ.get("PASSWORD")

# MAINTENANCE SWITCH & TRACKER
MAINTENANCE_MODE = True  
MAINTENANCE_ALERT_SENT = True

# ALLOWED TECHNICIANS
ALLOWED_TECHNICIANS = [
     "Girmaye Kelil", "Israel Aklilu",
     "Yared Girma", "Yohanis Getiye",
]

# GOOGLE FORM URL
FORM_URL = os.environ.get("GOOGLE_FORM_URL", "https://docs.google.com/forms/d/e/1FAIpQLSfJAWo1l6gNT2hFwnGZcf-ibX-8drfZLR_ww6JMx_yFZCEcGQ/formResponse")

# REAL & UPDATED GOOGLE FORM ENTRY IDs
ENTRY_EMAIL = "entry.111111111"             
ENTRY_TECH_NAME = "entry.206490333"         # Technician Name
ENTRY_BANK = "entry.2128913998"             # Bank
ENTRY_BRANCH = "entry.1983056024"           # Branch
ENTRY_DISTRICT = "entry.1132223049"         # District
ENTRY_TYPE2 = "entry.1173614214"            # pm / case

# CASE Specific Entry IDs
ENTRY_CASE_ID = "entry.283120155"           # Case Id
ENTRY_TERMINAL_NO = "entry.1541091566"      # Terminal No
ENTRY_CASE_ISSUE = "entry.1741675200"       # Case Issue
ENTRY_REG_TYPE = "entry.1717551465"         # Case Registration Type (Dashboard / Telegram)
ENTRY_CASE_TYPE = "entry.1287114682"        # Type (phone / physical)
ENTRY_STATUS = "entry.1994644026"           # Status (Completed / On going)
ENTRY_SPARE_PART = "entry.106596101"        # Spare Part (Yes / No)
ENTRY_PART_NAME = "entry.1167440013"        # Part Name
ENTRY_COMMENT = "entry.38555627"           # Comment

# Date & Time Fields
ENTRY_REG_DATE = "entry.2081498177"
ENTRY_REG_TIME = "entry.1802377317"
ENTRY_CLOSED_DATE = "entry.1340570535"
ENTRY_CLOSED_TIME = "entry.1091544422"


def get_eat_now():
    return datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(hours=3)

raw_chat_id = os.environ.get("NOTIFICATION_CHAT_ID", "")
if raw_chat_id.startswith("-") or raw_chat_id.isdigit():
    try: NOTIFICATION_CHAT_ID = int(raw_chat_id)
    except ValueError: NOTIFICATION_CHAT_ID = raw_chat_id
else:
    NOTIFICATION_CHAT_ID = raw_chat_id if raw_chat_id else None

SENT_CASES_TRACKER = set()
SENT_REMINDERS_TRACKER = {}
ACTIVE_USERS_TRACKER = set()

# MULTI-STEP FORM STATE TRACKER
USER_FORM_STATES = {}

# GLOBAL HTTP CLIENT
HTTP_CLIENT = httpx.AsyncClient(
    headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json',
        'Origin': 'https://tech24et.com',
        'Referer': 'https://tech24et.com/'
    },
    follow_redirects=True,
    timeout=15.0,
    verify=False
)

# ==========================================
# 2. FLASK SERVER FOR KEEPALIVE (RENDER)
# ==========================================
app = Flask(__name__)
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def catch_all(path):
    return "Bot is Running and Alive!", 200

def run_health_server():
    port = int(os.environ.get("PORT", 10000))
    logger.info(f"Starting Keep-Alive Flask server on port {port}...")
    app.run(host="0.0.0.0", port=port, use_reloader=False, threaded=True)

# ==========================================
# 3. ROBUST JSON FIELD EXTRACTORS
# ==========================================
def safe_parse_json(val):
    if not val: return {}
    if isinstance(val, dict): return val
    try:
        if isinstance(val, str):
            cleaned = val.replace("'", '"').replace("None", "null").replace("True", "true").replace("False", "false")
            return json.loads(cleaned)
    except Exception: pass
    return {}

def clean_extracted_value(data, key_hierarchy):
    if not data: return ""
    parsed_data = safe_parse_json(data) if isinstance(data, str) else data
    if not isinstance(parsed_data, dict): return str(parsed_data)
    
    for key in key_hierarchy:
        if key in parsed_data and parsed_data[key] is not None:
            val = parsed_data[key]
            if isinstance(val, dict): 
                return clean_extracted_value(val, key_hierarchy)
            return str(val)
            
    for key in key_hierarchy:
        for k, v in parsed_data.items():
            if isinstance(v, dict) and key in v:
                return str(v[key])
    return ""

def get_relative_time(date_obj):
    if not date_obj: return "Just now", "Just now"
    now = get_eat_now()
    diff = now - date_obj
    seconds = diff.total_seconds()
    if seconds < 0:
        minutes = abs(int(seconds // 60))
        if minutes < 2: return "Just now", "Just now"
        return f"{minutes}min", f"{minutes}min"
    minutes = int(seconds // 60)
    hours = int(minutes // 60)
    days = int(hours // 24)
    if days > 0: time_str = f"{days}d"
    elif hours > 0: time_str = f"{hours}h"
    else: time_str = f"{minutes}min"
    return time_str, time_str

def find_matching_technician(dashboard_tech_name):
    if not dashboard_tech_name or str(dashboard_tech_name).strip().lower() in ["none", "not assigned", "-", "null"]:
        return None
    dash_clean = " ".join(str(dashboard_tech_name).strip().split()).lower()
    for tech in ALLOWED_TECHNICIANS:
        tech_clean = " ".join(str(tech).strip().split()).lower()
        if tech_clean == dash_clean: return tech
    return None

# ==========================================
# 4. API SCRAPER
# ==========================================
async def scrape_website_cases():
    if not EMAIL or not PASSWORD:
        return [], "Error: EMAIL or PASSWORD environment variables missing!"

    csrf_url = 'https://api.tech24et.com/sanctum/csrf-cookie'
    login_url = 'https://api.tech24et.com/api/login'
    api_url = 'https://api.tech24et.com/api/callentries?limit=250'

    try:
        await HTTP_CLIENT.get(csrf_url)
        xsrf_token = HTTP_CLIENT.cookies.get("XSRF-TOKEN")
        if xsrf_token:
            HTTP_CLIENT.headers.update({'X-XSRF-TOKEN': urllib.parse.unquote(xsrf_token)})

        payload = {'email': EMAIL.strip(), 'password': PASSWORD.strip()}
        login_res = await HTTP_CLIENT.post(login_url, json=payload)
        if login_res.status_code not in [200, 201, 204]:
            return [], f"Login failed! Code: {login_res.status_code}"

        response = await HTTP_CLIENT.get(api_url)
        if response.status_code != 200:
            return [], f"API GET error: {response.status_code}"

        data = response.json()
        raw_list = data.get('data', []) if isinstance(data, dict) else data
        if not isinstance(raw_list, list):
            return [], "Error: Data response format isn't parsed into list."

        scraped_cases = []
        for entry in raw_list:
            if not entry or not isinstance(entry, dict): continue
            
            raw_string_dump = str(entry).lower()
            if "adama" in raw_string_dump:
                case_id = str(entry.get('callentry_id', 'N/A'))
                
                terminal_data = entry.get('atmterminal') or {}
                terminal_no = clean_extracted_value(terminal_data, ['atmterminal_no', 'terminal_no']) if isinstance(terminal_data, dict) else 'N/A'
                if not terminal_no or terminal_no == "None": terminal_no = "N/A"
                
                terminal_name = clean_extracted_value(terminal_data, ['atmterminal_name', 'terminal_name']) if isinstance(terminal_data, dict) else 'N/A'
                if not terminal_name or terminal_name == "None": terminal_name = "N/A"

                bank_data = entry.get('bank') or {}
                bank = clean_extracted_value(bank_data, ['bank_name', 'bankname']) if isinstance(bank_data, dict) else 'Awash'
                if not bank or bank == "None": bank = "Awash"

                issue_data = entry.get('issuesubcategory') or entry.get('issuecategory') or {}
                issue = clean_extracted_value(issue_data, ['issuesubcat_name', 'issuecatname', 'name']) if isinstance(issue_data, dict) else 'ATM Issue'
                if not issue or issue == "None": issue = "ATM Issue"

                branch_data = entry.get('branch') or {}
                branch = clean_extracted_value(branch_data, ['branch_name', 'branchname']) if isinstance(branch_data, dict) else 'Adama Branch'
                if not branch or branch == "None": branch = "Adama Branch"

                district_data = entry.get('district') or {}
                district = clean_extracted_value(district_data, ['dist_name', 'district_name']) if isinstance(district_data, dict) else 'Adama'
                if not district or district == "None": district = "Adama"

                comment = entry.get('callentry_description') or "-"
                if not comment or comment.strip() == "": comment = "-"

                technician = entry.get('assigned_eng', 'Not Assigned')
                if not technician or str(technician).strip() == "" or str(technician).lower() == "none":
                    tech_obj = entry.get('Technician') or entry.get('technician') or {}
                    if isinstance(tech_obj, dict):
                        technician = tech_obj.get('assigned_eng', 'Not Assigned')

                if not technician or str(technician).strip() == "" or str(technician).lower() == "none":
                    technician = "Not Assigned"
                
                tech_phone = entry.get('assigned_phone', '-')
                if not tech_phone: tech_phone = "-"

                created_at = entry.get('created_at') or entry.get('Reported At') or entry.get('updated_at')
                closed_at_raw = entry.get('closed_at') or entry.get('updated_at') or ""

                date_obj = None
                if created_at:
                    date_str = str(created_at).strip()
                    formats_to_try = (
                        "%Y-%m-%d %I:%M %p", "%Y-%m-%d %I:%M:%S %p",
                        "%d/%m/%Y %I:%M %p", "%d/%m/%Y %I:%M:%S %p",
                        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                        "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", 
                        "%d/%m/%Y", "%Y-%m-%d"
                    )
                    clean_time_str = date_str.split(".")[0].replace("T", " ")
                    for fmt in formats_to_try:
                        try:
                            date_obj = datetime.strptime(clean_time_str, fmt).replace(tzinfo=None)
                            break
                        except ValueError: continue
                            
                if not date_obj:
                    date_obj = get_eat_now()

                reg_date = date_obj.strftime("%d/%m/%Y")
                reg_time = date_obj.strftime("%I:%M %p")
                date_str = f"{reg_date} {reg_time}"

                closed_date, closed_time = ("-", "-")
                closed_date_obj = None
                if closed_at_raw and " " in str(closed_at_raw):
                    c_str = str(closed_at_raw).split(".")[0].replace("T", " ")
                    for fmt in (
                        "%Y-%m-%d %I:%M %p", "%d/%m/%Y %I:%M %p",
                        "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", 
                        "%d/%m/%Y", "%Y-%m-%d"
                    ):
                        try:
                            closed_date_obj = datetime.strptime(c_str, fmt).replace(tzinfo=None)
                            break
                        except ValueError: pass
                    if closed_date_obj:
                        closed_date = closed_date_obj.strftime("%d/%m/%Y")
                        closed_time = closed_date_obj.strftime("%I:%M %p")

                status_raw = str(entry.get('callentry_status', '')).lower()
                if status_raw in ["complete", "completed", "done", "1"]: status_text = "Completed"
                else: status_text = "On going"

                scraped_cases.append({
                    'case_id': case_id, 'bank': bank, 'district': district, 'branch': branch,
                    'terminal': terminal_no, 'atm_name': terminal_name, 'issue': issue,
                    'status': status_text, 'comment': comment, 'technician': technician,
                    'tech_phone': tech_phone, 'date_raw': date_str, 'date_obj': date_obj,
                    'closed_date_obj': closed_date_obj,
                    'reg_date': reg_date, 'reg_time': reg_time,
                    'closed_date': closed_date, 'closed_time': closed_time
                })
        return scraped_cases, "OK"
    except Exception as e:
        logger.error(f"Scraper Exception: {str(e)}")
        return [], f"Scraper Exception: {str(e)}"

# ==========================================
# TERMINATE CASE FUNCTION
# ==========================================
async def terminate_case_on_dashboard(case_id):
    terminate_url = f'https://api.tech24et.com/api/callentries/{case_id}/close'
    update_url = f'https://api.tech24et.com/api/callentries/{case_id}'
    login_url = "https://api.tech24et.com/api/login"
    csrf_url = 'https://api.tech24et.com/sanctum/csrf-cookie'

    try:
        HTTP_CLIENT.cookies.clear()
        await HTTP_CLIENT.get(csrf_url)
        xsrf_token = HTTP_CLIENT.cookies.get("XSRF-TOKEN")
        if xsrf_token:
            HTTP_CLIENT.headers.update({
                'X-XSRF-TOKEN': urllib.parse.unquote(xsrf_token),
                'Referer': 'https://tech24et.com/'
            })

        payload = {"email": EMAIL.strip(), "password": PASSWORD.strip()}
        login_res = await HTTP_CLIENT.post(login_url, json=payload)
        if login_res.status_code not in [200, 201, 204]:
            return False, f"Auth Error status: {login_res.status_code}"

        fresh_xsrf = HTTP_CLIENT.cookies.get("XSRF-TOKEN")
        if fresh_xsrf:
            HTTP_CLIENT.headers.update({'X-XSRF-TOKEN': urllib.parse.unquote(fresh_xsrf)})

        res = await HTTP_CLIENT.post(terminate_url, json={"status": "Completed", "callentry_status": "Completed"})
        if res.status_code in [200, 201, 204]:
            return True, "Successfully Closed"

        res_put = await HTTP_CLIENT.put(update_url, json={"status": "Completed", "callentry_status": "Completed"})
        if res_put.status_code in [200, 201, 204]:
            return True, "Successfully Closed"

        return False, f"API Rejected: Code {res.status_code}"
    except Exception as e:
        logger.error(f"Terminate Case Exception: {str(e)}")
        return False, str(e)

# ==========================================
# 5. AUTOMATIC ALARM & OVERDUE LOOP
# ==========================================
async def check_and_alert_cases(bot, target_user_id=None):
    cases, status = await scrape_website_cases()
    if status != "OK":
        logger.error(f"Alert Engine Scraper error: {status}")
        return

    pending_statuses = ["on going", "pending", "open", "0"]
    pending_cases = [c for c in cases if str(c.get('status', '')).lower() in pending_statuses or c.get('status') == "On going"]
    now = get_eat_now()

    for case in pending_cases:
        case_id = case['case_id']
        case_time = case.get('date_obj', now)

        time_diff = now - case_time
        hours_ago = int(time_diff.total_seconds() // 3600)
        mins_ago = int((time_diff.total_seconds() % 3600) // 60)
        age_str = f"{hours_ago}h {mins_ago}m ago" if hours_ago > 0 else f"{mins_ago}min ago"

        notif_text = (
            f"🚨 *ATM Incident Alert* 🚨\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📄 *ID:* `{case_id}`\n"
            f"🏦 *Bank:* {case['bank']}\n"
            f"🏢 *Branch:* {case['branch']}\n"
            f"⚠️ *Issue:* {case['issue']}\n"
            f"📍 *District:* {case['district']}\n"
            f"👤 *Technician:* {case['technician']}\n"
            f"💬 *Comment:* {case['comment']}\n"
            f"🕒 *Reported at:* {case['date_raw']} ({age_str})\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📌 _Status: Pending Action / Unresolved_"
        )
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("Check in dashboard", url="https://tech24et.com/login")]])
        
        if target_user_id:
            try:
                await bot.send_message(chat_id=target_user_id, text=notif_text, reply_markup=kb, parse_mode="Markdown")
            except Exception as e:
                logger.error(f"Failed to send startup alert to user {target_user_id}: {str(e)}")
            if case_id not in SENT_CASES_TRACKER:
                SENT_CASES_TRACKER.add(case_id)
            continue

        if case_id not in SENT_CASES_TRACKER:
            SENT_CASES_TRACKER.add(case_id)
            
            if NOTIFICATION_CHAT_ID:
                try: 
                    await bot.send_message(chat_id=NOTIFICATION_CHAT_ID, text=notif_text, reply_markup=kb, parse_mode="Markdown")
                except Exception as e: 
                    logger.warning(f"Could not send to NOTIFICATION_CHAT_ID: {str(e)}")

            for user_id in list(ACTIVE_USERS_TRACKER):
                try: 
                    await bot.send_message(chat_id=user_id, text=notif_text, reply_markup=kb, parse_mode="Markdown")
                except Exception: 
                    pass
            continue

        time_elapsed = now - case_time
        if time_elapsed >= timedelta(hours=5):
            last_reminder = SENT_REMINDERS_TRACKER.get(case_id)
            if last_reminder is None or (now - last_reminder) >= timedelta(hours=5):
                SENT_REMINDERS_TRACKER[case_id] = now
                hours_passed = int(time_elapsed.total_seconds() // 3600)
                
                reminder_text = (
                    f"⚠️ *OVERDUE INCIDENT REMINDER (>{hours_passed} Hours)* ⚠️\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"This case has been active for more than {hours_passed} hours. Please investigate.\n\n"
                    f"📄 *ID:* `{case_id}`\n"
                    f"🏦 *Bank:* {case['bank']} ({case['branch']})\n"
                    f"⚠️ *Issue:* {case['issue']}\n"
                    f"👤 *Technician:* {case['technician']}\n"
                    f"🕒 *Reported at:* {case['date_raw']}\n\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"⏳ _Duration: Still Pending!_"
                )
                
                if NOTIFICATION_CHAT_ID:
                    try: 
                        await bot.send_message(chat_id=NOTIFICATION_CHAT_ID, text=reminder_text, reply_markup=kb, parse_mode="Markdown")
                    except Exception as e: 
                        logger.error(f"Failed to send overdue reminder to NOTIFICATION_CHAT_ID: {str(e)}")
                
                for user_id in list(ACTIVE_USERS_TRACKER):
                    try: 
                        await bot.send_message(chat_id=user_id, text=reminder_text, reply_markup=kb, parse_mode="Markdown")
                    except Exception as e: 
                        logger.warning(f"Failed to send overdue reminder to individual user {user_id}: {str(e)}")

# 💡 💡 MAINTENANCE ALERT SENDER LOOP 💡 💡
async def start_independent_alarm_loop(bot):
    global MAINTENANCE_ALERT_SENT
    logger.info("Background Alarm Engine successfully launched inside Application Loop.")
    
    while True:
        try:
            if MAINTENANCE_MODE:
                if not MAINTENANCE_ALERT_SENT:
                    m_msg = get_maintenance_message()
                    
                    if NOTIFICATION_CHAT_ID:
                        try:
                            await bot.send_message(chat_id=NOTIFICATION_CHAT_ID, text=m_msg, parse_mode="Markdown")
                        except Exception as e:
                            logger.error(f"Failed to send maintenance alert to NOTIFICATION_CHAT_ID: {e}")

                    for user_id in list(ACTIVE_USERS_TRACKER):
                        try:
                            await bot.send_message(chat_id=user_id, text=m_msg, parse_mode="Markdown")
                        except Exception as e:
                            logger.warning(f"Failed to send maintenance alert to user {user_id}: {e}")

                    MAINTENANCE_ALERT_SENT = True
            else:
                MAINTENANCE_ALERT_SENT = False
                await check_and_alert_cases(bot)
                
        except Exception as e:
            logger.error(f"Error inside independent background loop: {str(e)}")
            
        await asyncio.sleep(20)

def get_maintenance_message():
    return (
        "🚨 *SYSTEM NOTICE / MAINTENANCE ALERT* 🚨\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "⚠️ *For All bot users!!!*\n"
        "The bot is currently under maintenance. We are performing system optimizations. Please be patient 🙏 Thank you for understanding.\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "🛠️ _Status: Upgrading systems & optimization ongoing_"
    )

def build_case_detail_ui(case):
    relative_long, _ = get_relative_time(case.get('date_obj'))
    text = (
        f"Case ID: {case['case_id']}\n"
        f"Terminal: {case['terminal']}\n"
        f"Bank: {case['bank']}\n"
        f"Branch: {case['branch']}\n"
        f"Issue: {case['issue']}\n"
        f"Status: {case['status']}\n"
        f"District: {case['district']}\n"
        f"Comment: {case['comment']}\n"
        f"Technician: {case['technician']}\n"
        f"Reported At: {case['date_raw']} (EAT)\n"
        f"Relative Time: {relative_long}"
    )
    keyboard = [
        [InlineKeyboardButton("⛔ Terminate", callback_data=f"askterm_{case['case_id']}")],
        [InlineKeyboardButton("🌀 Refresh", callback_data=f"refresh_{case['case_id']}")],
        [InlineKeyboardButton("⛔ Cancel", callback_data="cancel_action")]
    ]
    return text, InlineKeyboardMarkup(keyboard)

# ==========================================
# 7. EXCEL & SPECIFIC REPORT FORMATTERS
# ==========================================
def format_technician_daily_report(cases, selected_tech, report_type):
    now = get_eat_now()
    today_str = now.strftime("%d/%m/%Y")
    filtered_cases = []
    for c in cases:
        if c.get('date_obj') and c['date_obj'].strftime("%d/%m/%Y") == today_str:
            matched_tech = find_matching_technician(c['technician'])
            if matched_tech and matched_tech.lower() == selected_tech.lower():
                filtered_cases.append(c)

    title_type = "Telegram Registered Cases" if report_type == "case" else "PM Report" if report_type == "pm" else "Dashboard Cases"
    if not filtered_cases:
        return (
            f"📋 *Adama District Daily Report ({title_type}) - {selected_tech}* 📋\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📭 *Currently, there are no recorded cases for this technician today ({today_str}).*\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
    report_lines = [f"📋 *Adama District Daily Report ({title_type}) - {selected_tech}* 📋\n", f"📅 Date: {today_str}\n━━━━━━━━━━━━━━━━━━━━━━━━━━"]
    for idx, c in enumerate(filtered_cases, start=1):
        status_emoji = "✅ Completed" if c['status'] == "Completed" else "⏳ On going"
        line = f"{idx}. ID: {c['case_id']}\n🏦 Bank: {c['bank']} ({c['branch']} branch)\n⚠️ Issue: {c['issue']}\n📌 Status: {status_emoji}\n💬 Comment: {c['comment']}\n----------------------------------------"
        report_lines.append(line)
    return "\n".join(report_lines)

def format_technician_weekly_report(cases, selected_tech):
    now = get_eat_now()
    days_since_sunday = (now.weekday() + 1) % 7
    start_of_week = (now - timedelta(days=days_since_sunday)).replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_week = start_of_week + timedelta(days=6, hours=23, minutes=59, seconds=59)

    filtered_cases = []
    for c in cases:
        c_date = c.get('date_obj')
        c_closed_date = c.get('closed_date_obj')
        
        is_created_this_week = c_date and (start_of_week <= c_date <= end_of_week)
        is_closed_this_week = c_closed_date and (start_of_week <= c_closed_date <= end_of_week)

        if is_created_this_week or is_closed_this_week:
            matched_tech = find_matching_technician(c['technician'])
            if matched_tech and matched_tech.lower() == selected_tech.lower():
                filtered_cases.append(c)

    if not filtered_cases:
        return (
            f"📋 *Adama District Weekly Cases Report - {selected_tech}* 📋\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📭 *Currently, there are no recorded cases assigned to this technician for this week.*\n\n"
            f"🌟 Keep up the great work!\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )

    report_lines = [f"📋 *Adama District Weekly Cases Report - {selected_tech}* 📋\n"]

    for idx, c in enumerate(filtered_cases, start=1):
        reg_datetime_str = c.get('date_raw', '-')
        
        if c.get('status') == "Completed":
            if c.get('closed_date') != "-" and c.get('closed_time') != "-":
                closed_str = f"{c['closed_date']} {c['closed_time']}"
            else:
                closed_str = "Completed (No timestamp)"
            status_line = f"📌 Status: ✅ Completed\n✅ Closed: {closed_str}"
        else:
            status_line = "📌 Status: ⏳ On going"

        line = (
            f"{idx}. ID: {c['case_id']}\n"
            f"🏦 Bank: {c['bank']} ({c['branch']} branch)\n"
            f"⚠️ Issue: {c['issue']}\n"
            f"📅 Reported: {reg_datetime_str}\n"
            f"{status_line}\n"
            f"----------------------------------------"
        )
        report_lines.append(line)

    report_lines.append("\n        *Summary Overview*")
    bank_analytics = {}
    for case in filtered_cases:
        b_name = case['bank']
        if b_name not in bank_analytics: 
            bank_analytics[b_name] = {"completed": 0, "ongoing": 0}
        if case['status'] == "Completed": 
            bank_analytics[b_name]["completed"] += 1
        else: 
            bank_analytics[b_name]["ongoing"] += 1

    for bank_name, stats in bank_analytics.items():
        report_lines.append(f"*{bank_name} Bank*\n    Completed: {stats['completed']}\n    On going: {stats['ongoing']}")
        
    return "\n".join(report_lines)

def format_weekly_summary_matrix(cases):
    now = get_eat_now()
    days_since_sunday = (now.weekday() + 1) % 7
    start_of_week = (now - timedelta(days=days_since_sunday)).replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_week = start_of_week + timedelta(days=6, hours=23, minutes=59, seconds=59)

    filtered_cases = [c for c in cases if c.get('date_obj') and (start_of_week <= c['date_obj'] <= end_of_week)]
    report_lines = ["📋 *Weekly Summary Report of Matrix* 📋\n"]

    tech_stats = {tech: {"completed": 0, "ongoing": 0} for tech in ALLOWED_TECHNICIANS}
    total_completed, total_ongoing, other_district_or_unassigned = 0, 0, 0

    for case in filtered_cases:
        matched_tech = find_matching_technician(case['technician'])
        if matched_tech:
            if case['status'] == "Completed":
                tech_stats[matched_tech]["completed"] += 1
                total_completed += 1
            else:
                tech_stats[matched_tech]["ongoing"] += 1
                total_ongoing += 1
        else: other_district_or_unassigned += 1

    for tech in ALLOWED_TECHNICIANS:
        stats = tech_stats[tech]
        report_lines.append(f" 👤 Technician *{tech}*: {stats['completed']} completed, {stats['ongoing']} ongoing.\n")

    report_lines.append(f"\n 🟧 Total in *Adama District*: {total_completed} completed, {total_ongoing} ongoing cases.")
    if other_district_or_unassigned > 0: report_lines.append(f" 🔍 Unassigned / Other District Cases: *{other_district_or_unassigned}*")
    total_cases = total_completed + total_ongoing
    if total_cases > 0: report_lines.append(f"🎯 Completion Rate: *{(total_completed / total_cases) * 100:.1f}%*")
    num_techs = len(ALLOWED_TECHNICIANS)
    if num_techs > 0: report_lines.append(f"📊 Average Completed cases per Tech: *{total_completed / num_techs:.1f}*")
    return "\n".join(report_lines)

def generate_excel_bytes(cases):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incident Log Database"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, bold=False, color="000000")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    headers = ["Case ID", "Terminal", "Bank", "Branch", "Issue Description", "Status", "District", "Comment", "Technician", "Tech Phone", "Date EAT"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, case in enumerate(cases, start=2):
        row_data = [case['case_id'], case['terminal'], case['bank'], case['branch'], case['issue'], case['status'], case['district'], case['comment'], case['technician'], case['tech_phone'], case['date_raw']]
        ws.append(row_data)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center") if col_num in [1, 2, 6, 10, 11] else Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# 8. TELEGRAM COMMAND HANDLERS
# ==========================================
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    ACTIVE_USERS_TRACKER.add(chat_id)
    if MAINTENANCE_MODE: return await update.message.reply_text(get_maintenance_message(), parse_mode="Markdown")

    welcome_text = (
        "👋 *Welcome to Tech24 Adama District Bot*\n\n"
        "💻 *Available Commands Menu:*\n"
        "• /pending - View currently open / unresolved cases\n"
        "• /daily - View daily report by technician selection\n"
        "• /report - View weekly performance metrics by technician\n"
        "• /summary - View overall weekly matrix summary\n"
        "• /export - Download structured incident Excel spreadsheets\n\n"
        "🔄 _Checking for live pending cases now..._"
    )
    await update.message.reply_text(welcome_text, parse_mode="Markdown")
    await check_and_alert_cases(context.bot, target_user_id=chat_id)

async def pending_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if MAINTENANCE_MODE: return await update.message.reply_text(get_maintenance_message(), parse_mode="Markdown")
    processing = await update.message.reply_text("⏳ Searching dashboard portal for Adama logs, please wait...")
    cases, status = await scrape_website_cases()
    try:
        await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=processing.message_id)
    except Exception: pass

    if status != "OK": return await update.message.reply_text(f"❌ *Connection Failure:*\n{status}", parse_mode="Markdown")
    pending_cases = [c for c in cases if c['status'] == "On going"]
    if not pending_cases:
        return await update.message.reply_text("✅ All Adama cases are completed! No pending cases found.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Check in dashboard", url="https://tech24et.com/login")]]))

    if len(pending_cases) == 1:
        text, kb = build_case_detail_ui(pending_cases[0])
        await update.message.reply_text(text, reply_markup=kb)
    else:
        text = "The following ATM cases have been reported and are currently pending action. Select a case from the list below to view details."
        keyboard = [[InlineKeyboardButton(f"{get_relative_time(c.get('date_obj'))[1]} | {c['case_id']} | {c['bank']} | {c['branch']}", callback_data=f"view_{c['case_id']}")] for c in pending_cases]
        keyboard.append([InlineKeyboardButton("Cancel", callback_data="cancel_action")])
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

async def daily_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if MAINTENANCE_MODE: return await update.message.reply_text(get_maintenance_message(), parse_mode="Markdown")
    keyboard = [[InlineKeyboardButton(f"👤 {tech}", callback_data=f"dtech_{tech}")] for tech in sorted(ALLOWED_TECHNICIANS)]
    keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel_action")])
    await update.message.reply_text("📋 *Daily Report Menu*\n\n 👥 Please select an Adama District Technician:", parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard))

async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if MAINTENANCE_MODE: return await update.message.reply_text(get_maintenance_message(), parse_mode="Markdown")
    keyboard = [[InlineKeyboardButton(f"👤 {tech}", callback_data=f"wrep_{tech}")] for tech in sorted(ALLOWED_TECHNICIANS)]
    keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel_action")])
    await update.message.reply_text("📊 *Weekly Report Menu*\n\n 👥 Select an Adama District Technician to view their weekly cases report (Sunday - Saturday):", parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard))

async def summary_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if MAINTENANCE_MODE: return await update.message.reply_text(get_maintenance_message(), parse_mode="Markdown")
    processing = await update.message.reply_text("⏳ Searching dashboard portal for Adama logs, please wait...")
    cases, status = await scrape_website_cases()
    try:
        await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=processing.message_id)
    except Exception: pass
    if status != "OK": return await update.message.reply_text(f"❌ *Error:* {status}", parse_mode="Markdown")
    await update.message.reply_text(format_weekly_summary_matrix(cases), parse_mode="Markdown")

async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if MAINTENANCE_MODE: return await update.message.reply_text(get_maintenance_message(), parse_mode="Markdown")
    processing = await update.message.reply_text("⏳ Writing and formatting Excel spreadsheet...")
    cases, status = await scrape_website_cases()
    try:
        await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=processing.message_id)
    except Exception: pass

    if status != "OK": return await update.message.reply_text(f"❌ *Export Blocked:* {status}", parse_mode="Markdown")
    if not cases: return await update.message.reply_text("❌ *Export Cancelled:* No cases matched query scope.", parse_mode="Markdown")

    excel_file = generate_excel_bytes(cases)
    excel_file.name = f"case-report-{get_eat_now().strftime('%Y-%m')}.xlsx"
    await context.bot.send_document(chat_id=update.effective_chat.id, document=excel_file, caption=f"📊 *ATM Cases Report – {get_eat_now().strftime('%B %Y')}*\n\nThis report contains all ATM cases.", parse_mode="Markdown")

def get_bank_selection_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🏦 Awash Bank", callback_data="fbank_Awash"),
         InlineKeyboardButton("🏦 Dashen Bank", callback_data="fbank_Dashen")],
        [InlineKeyboardButton("🏦 Ahadu Bank", callback_data="fbank_Ahadu"),
         InlineKeyboardButton("➕ Other Bank", callback_data="fbank_Other")],
        [InlineKeyboardButton("❌ Cancel Process", callback_data="cancel_action")]
    ])

# ==========================================
# 9. INLINE BUTTON CALLBACK HANDLER
# ==========================================
async def button_click_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = update.effective_chat.id

    if MAINTENANCE_MODE:
        await context.bot.send_message(chat_id=chat_id, text=get_maintenance_message(), parse_mode="Markdown")
        return

    data = query.data
    if data == "cancel_action":
        USER_FORM_STATES.pop(chat_id, None)
        try:
            await query.message.delete()
        except Exception: pass
        return

    if data.startswith("dtech_"):
        tech_name = data.split("_")[1]
        confirm_text = (
            f"🔥 *Daily Report Option Menu for {tech_name}*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"ℹ️ *For Dashboard Cases Report:* 👉 Press the *📊 Dashboard* button.\n\n"
            f"ℹ️ *For Telegram Cases & PM Reports:* 👉 Press the *📱 Telegram & PM* button.\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
        confirm_keyboard = [
            [InlineKeyboardButton("📊 Dashboard", callback_data=f"ddash_{tech_name}"),
             InlineKeyboardButton("📱 Telegram & PM", callback_data=f"dtgpm_menu_{tech_name}")],
            [InlineKeyboardButton("🔙 Back to Technicians", callback_data="back_to_daily_techs")]
        ]
        await query.edit_message_text(text=confirm_text, reply_markup=InlineKeyboardMarkup(confirm_keyboard), parse_mode="Markdown")
        return

    if data.startswith("dtgpm_menu_"):
        tech_name = data.split("_")[2]
        tgpm_text = (
            f"📱 *Telegram & PM Report Options ({tech_name})*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"ℹ️ *To register cases received via Telegram:* 👉 Press *📋 CASE*.\n\n"
            f"ℹ️ *To submit a PM (Preventive Maintenance) report:* 👉 Press *⚙️ PM*.\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
        tgpm_keyboard = [
            [InlineKeyboardButton("📋 CASE", callback_data=f"drpt_case_{tech_name}"),
             InlineKeyboardButton("⚙️ PM", callback_data=f"drpt_pm_{tech_name}")],
            [InlineKeyboardButton("🔙 Back", callback_data=f"dtech_{tech_name}")]
        ]
        await query.edit_message_text(text=tgpm_text, reply_markup=InlineKeyboardMarkup(tgpm_keyboard), parse_mode="Markdown")
        return

    if data.startswith("drpt_case_"):
        tech_name = data.split("_")[2]
        now_eat = get_eat_now()
        
        USER_FORM_STATES[chat_id] = {
            'step': 'SELECT_BANK_NAME',
            'tech_name': tech_name,
            'extracted_payload': {
                ENTRY_TECH_NAME: tech_name,
                ENTRY_TYPE2: 'case',
                ENTRY_REG_TYPE: 'Telegram',
                ENTRY_DISTRICT: 'Adama',
                ENTRY_REG_DATE: now_eat.strftime("%d/%m/%Y"),
                ENTRY_REG_TIME: now_eat.strftime("%I:%M %p")
            }
        }
        
        await query.edit_message_text(
            text=f"📋 *New Case Registration Form ({tech_name})*\n📅 Date: `{now_eat.strftime('%d/%m/%Y %I:%M %p')}` (Auto-inserted)\n\n🏦 *Please select the Bank Name:*", 
            reply_markup=get_bank_selection_keyboard(),
            parse_mode="Markdown"
        )
        return

    if data.startswith("drpt_pm_"):
        tech_name = data.split("_")[2]
        now_eat = get_eat_now()
        
        USER_FORM_STATES[chat_id] = {
            'step': 'PM_SELECT_BANK_NAME',
            'tech_name': tech_name,
            'extracted_payload': {
                ENTRY_TECH_NAME: tech_name,
                ENTRY_TYPE2: 'pm',
                ENTRY_STATUS: 'Completed',
                ENTRY_DISTRICT: 'Adama',
                ENTRY_REG_DATE: now_eat.strftime("%d/%m/%Y"),
                ENTRY_REG_TIME: now_eat.strftime("%I:%M %p")
            }
        }
        
        await query.edit_message_text(
            text=f"⚙️ *New PM (Preventive Maintenance) Form ({tech_name})*\n📅 Date: `{now_eat.strftime('%d/%m/%Y %I:%M %p')}` (Auto-inserted)\n📌 Status: `Completed` (Auto-inserted)\n\n🏦 *Please select the Bank Name:*", 
            reply_markup=get_bank_selection_keyboard(),
            parse_mode="Markdown"
        )
        return

    if data.startswith("fbank_"):
        selected_bank = data.split("fbank_")[1]
        if chat_id not in USER_FORM_STATES: return

        if selected_bank == "Other":
            USER_FORM_STATES[chat_id]['step'] = 'WAITING_FOR_CUSTOM_BANK_NAME'
            await query.edit_message_text("✍️ *Please type the Bank Name:*", parse_mode="Markdown")
            return

        USER_FORM_STATES[chat_id]['extracted_payload'][ENTRY_BANK] = f"{selected_bank} Bank" if not selected_bank.endswith("Bank") else selected_bank
        
        if USER_FORM_STATES[chat_id]['extracted_payload'].get(ENTRY_TYPE2) == 'pm':
            USER_FORM_STATES[chat_id]['step'] = 'PM_WAITING_FOR_BRANCH_NAME'
        else:
            USER_FORM_STATES[chat_id]['step'] = 'WAITING_FOR_BRANCH_NAME'

        await query.edit_message_text("🏢 *Please enter the Branch Name:*", parse_mode="Markdown")
        return

    if data.startswith("ddash_"):
        tech_name = data.split("_")[1]
        await query.edit_message_text("⏳ Syncing logs from dashboard portal...")
        cases, status = await scrape_website_cases()
        if status != "OK": 
            return await query.edit_message_text(f"❌ API Sync Fail: {status}")

        now = get_eat_now()
        today_str = now.strftime("%d/%m/%Y")

        filtered_cases = []
        for c in cases:
            matched_tech = find_matching_technician(c['technician'])
            if matched_tech and matched_tech.lower() == tech_name.lower():
                case_status = str(c.get('status', '')).lower()
                c_date_str = c.get('date_obj').strftime("%d/%m/%Y") if c.get('date_obj') else c.get('reg_date')
                c_closed_str = c.get('closed_date_obj').strftime("%d/%m/%Y") if c.get('closed_date_obj') else c.get('closed_date')

                if case_status in ["on going", "pending", "open"]:
                    filtered_cases.append(c)
                elif case_status in ["completed", "done"]:
                    if c_date_str == today_str or c_closed_str == today_str:
                        filtered_cases.append(c)

        if not filtered_cases:
            return await query.edit_message_text(
                text=f"📭 *No active pending cases or today's completed cases found for {tech_name}.*", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data=f"dtech_{tech_name}")]])
            )

        text = f"📊 *Dashboard Cases for {tech_name}:*\n_(Showing all Pending cases + Today's Completed cases)_"
        
        keyboard = []
        for c in filtered_cases:
            status_icon = "✅" if c['status'] == "Completed" else "⏳"
            btn_text = f"{status_icon} ID: {c['case_id']} | {c['branch']} ({c['status']})"
            keyboard.append([InlineKeyboardButton(btn_text, callback_data=f"fcase_{c['case_id']}")])

        keyboard.append([InlineKeyboardButton("🔙 Back", callback_data=f"dtech_{tech_name}")])
        await query.edit_message_text(text=text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")
        return

    if data.startswith("fcase_"):
        case_id = data.split("_")[1]
        await query.edit_message_text("⏳ Extracting data for Google Form mapping...")
        cases, status = await scrape_website_cases()
        if status != "OK": return await query.edit_message_text(f"❌ Sync Fail: {status}")
        
        target_case = next((c for c in cases if str(c['case_id']) == str(case_id)), None)
        if not target_case: return await query.edit_message_text("❌ Selected record could not be found.")

        USER_FORM_STATES[chat_id] = {
            'step': 'ASK_TYPE',
            'tech_name': target_case['technician'],
            'extracted_payload': {
                ENTRY_TECH_NAME: target_case['technician'],
                ENTRY_CASE_ID: target_case['case_id'],
                ENTRY_TERMINAL_NO: target_case['terminal'],
                ENTRY_BANK: target_case['bank'],
                ENTRY_BRANCH: target_case['branch'],
                ENTRY_CASE_ISSUE: target_case['issue'],
                ENTRY_COMMENT: target_case['comment'],
                ENTRY_DISTRICT: target_case['district'],
                ENTRY_REG_DATE: target_case['reg_date'],
                ENTRY_REG_TIME: target_case['reg_time'],
                ENTRY_CLOSED_DATE: target_case['closed_date'],
                ENTRY_CLOSED_TIME: target_case['closed_time'],
                
                ENTRY_TYPE2: 'case',
                ENTRY_REG_TYPE: 'Dashboard',
            }
        }

        type_kb = [
            [InlineKeyboardButton("📱 Phone", callback_data="ftype_phone"),
             InlineKeyboardButton("🏃 Physical", callback_data="ftype_physical")],
            [InlineKeyboardButton("❌ Cancel Process", callback_data="cancel_action")]
        ]
        await context.bot.send_message(
            chat_id=chat_id, 
            text=f"📊 *Form Configurator Loaded for Case {case_id}*\n\n*1. Select Support Type:*", 
            reply_markup=InlineKeyboardMarkup(type_kb), 
            parse_mode="Markdown"
        )
        try:
            await query.message.delete()
        except Exception: pass
        return

    if data.startswith("ftype_"):
        selected_type = data.split("_")[1]
        if chat_id not in USER_FORM_STATES: return
        
        USER_FORM_STATES[chat_id]['extracted_payload'][ENTRY_CASE_TYPE] = selected_type
        USER_FORM_STATES[chat_id]['step'] = 'ASK_ISSUE'

        if ENTRY_CASE_ISSUE in USER_FORM_STATES[chat_id]['extracted_payload']:
            USER_FORM_STATES[chat_id]['step'] = 'ASK_STATUS'
            status_kb = [
                [InlineKeyboardButton("✅ Completed", callback_data="fstat_Completed"),
                 InlineKeyboardButton("⏳ Pending / On going", callback_data="fstat_On going")],
                [InlineKeyboardButton("❌ Cancel Process", callback_data="cancel_action")]
            ]
            await query.edit_message_text("📌 *Select Case Status:*", reply_markup=InlineKeyboardMarkup(status_kb), parse_mode="Markdown")
        else:
            await query.edit_message_text("⚠️ *Please enter the Case Issue description:*", parse_mode="Markdown")
        return

    if data.startswith("fstat_"):
        selected_status = data.split("fstat_")[1]
        if chat_id not in USER_FORM_STATES: return

        USER_FORM_STATES[chat_id]['extracted_payload'][ENTRY_STATUS] = selected_status
        USER_FORM_STATES[chat_id]['step'] = 'ASK_SPARE'

        spare_kb = [
            [InlineKeyboardButton("Yes", callback_data="fspare_Yes"),
             InlineKeyboardButton("No", callback_data="fspare_No")],
            [InlineKeyboardButton("❌ Cancel Process", callback_data="cancel_action")]
        ]
        await query.edit_message_text(
            text="*Was any Spare Part used?*", 
            reply_markup=InlineKeyboardMarkup(spare_kb), 
            parse_mode="Markdown"
        )
        return

    if data.startswith("fspare_"):
        has_spare = data.split("_")[1]
        if chat_id not in USER_FORM_STATES: return

        USER_FORM_STATES[chat_id]['extracted_payload'][ENTRY_SPARE_PART] = has_spare

        if has_spare == "Yes":
            USER_FORM_STATES[chat_id]['step'] = 'WAITING_FOR_PART_NAME'
            await query.edit_message_text(
                text="🔩 *Please type the name of the Part used:*", 
                parse_mode="Markdown"
            )
        else:
            USER_FORM_STATES[chat_id]['extracted_payload'][ENTRY_PART_NAME] = "None"
            USER_FORM_STATES[chat_id]['step'] = 'PREVIEW_READY'
            await render_summary_and_confirm(query.message, USER_FORM_STATES[chat_id])
        return

    if data == "f_final_submit":
        if chat_id not in USER_FORM_STATES: return
        await query.edit_message_text("🚀 Submitting data to Google Forms...")
        
        payload = USER_FORM_STATES[chat_id]['extracted_payload']
        try:
            async with httpx.AsyncClient(timeout=20.0, follow_redirects=True) as client:
                resp = await client.post(FORM_URL, data=payload)
                if resp.status_code in [200, 302]:
                    await query.edit_message_text("✅ *Google Form Successfully Submitted!*", parse_mode="Markdown")
                else:
                    await query.edit_message_text(f"❌ *Submission Failed.* Status code: {resp.status_code}")
        except Exception as e:
            await query.edit_message_text(f"❌ *Network Error:* {str(e)}")
        
        USER_FORM_STATES.pop(chat_id, None)
        return

    if data == "back_to_daily_techs":
        keyboard = [[InlineKeyboardButton(f"👤 {tech}", callback_data=f"dtech_{tech}")] for tech in sorted(ALLOWED_TECHNICIANS)]
        keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel_action")])
        await query.edit_message_text("📋 *Daily Report Menu*\n\n 👥 Select an Adama District Technician:", parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard))
        return

    if data.startswith("wrep_"):
        tech_name = data.split("_")[1]
        cases, _ = await scrape_website_cases()
        await query.edit_message_text(
            text=format_technician_weekly_report(cases, tech_name),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Back to List", callback_data="back_to_techs")]
            ]),
            parse_mode="Markdown"
        )
        return

    if data == "back_to_techs":
        keyboard = [[InlineKeyboardButton(f"👤 {tech}", callback_data=f"wrep_{tech}")] for tech in sorted(ALLOWED_TECHNICIANS)]
        keyboard.append([InlineKeyboardButton("❌ Cancel", callback_data="cancel_action")])
        await query.edit_message_text("📊 *Weekly Report Menu*\n\n 👥 Select an Adama District Technician to view their weekly cases report (Sunday - Saturday):", parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard))
        return

    if data.startswith("askterm_"):
        case_id = data.split("_")[1]
        await query.edit_message_text(
            text=f"⚠️ *Confirmation Required*\n\nAre you sure you want to terminate/close Case ID: *{case_id}*?", 
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Go Back", callback_data=f"view_{case_id}")], 
                [InlineKeyboardButton("✅ Yes, Terminate", callback_data=f"do_terminate_{case_id}")], 
                [InlineKeyboardButton("❌ No, Cancel", callback_data="cancel_action")]
            ]), 
            parse_mode="Markdown"
        )
        return

    if data.startswith("do_terminate_"):
        case_id = data.split("_")[2]
        await query.edit_message_text(f"⏳ Attempting terminal closure for Case ID `{case_id}`...", parse_mode="Markdown")
        success, err_msg = await terminate_case_on_dashboard(case_id)
        if success: await query.edit_message_text(f"✅ *Success!* Case ID `{case_id}` marked as Terminated.", parse_mode="Markdown")
        else: await query.edit_message_text(text=f"❌ *Termination failed:*\n`{err_msg}`", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔄 Try Again", callback_data=f"askterm_{case_id}")], [InlineKeyboardButton("Cancel", callback_data="cancel_action")]]), parse_mode="Markdown")
        return

    if data.startswith("view_") or data.startswith("refresh_"):
        case_id = data.split("view_")[-1] if data.startswith("view_") else data.split("refresh_")[-1]
        cases, status = await scrape_website_cases()
        if status != "OK":
            return await query.edit_message_text(f"❌ Failed to fetch case details: {status}")
            
        target = next((c for c in cases if str(c['case_id']) == str(case_id)), None)
        if not target: 
            return await query.edit_message_text("❌ Record lost, unavailable, or finalized.")
            
        text, kb = build_case_detail_ui(target)
        await query.edit_message_text(text, reply_markup=kb)

async def render_summary_and_confirm(target_message, state_data):
    payload = state_data['extracted_payload']
    tech = state_data.get('tech_name', 'N/A')
    
    summary_msg = (
        f"📋 *Form Submission Review ({tech})* 📋\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📅 Date/Time: {payload.get(ENTRY_REG_DATE, '-')} {payload.get(ENTRY_REG_TIME, '-')}\n"
        f"🏷️ Registration Type: {payload.get(ENTRY_REG_TYPE, 'N/A')}\n"
        f"🏦 Bank Name: {payload.get(ENTRY_BANK, '-')}\n"
        f"🏢 Branch Name: {payload.get(ENTRY_BRANCH, '-')}\n"
    )
    
    if payload.get(ENTRY_TYPE2) == 'case':
        summary_msg += (
            f"📞 Support Type: {payload.get(ENTRY_CASE_TYPE, '-')}\n"
            f"⚠️ Issue: {payload.get(ENTRY_CASE_ISSUE, '-')}\n"
            f"🔩 Spare Used: {payload.get(ENTRY_SPARE_PART, '-')}\n"
            f"🏷️ Part Name: {payload.get(ENTRY_PART_NAME, '-')}\n"
            f"📌 Status: {payload.get(ENTRY_STATUS, '-')}\n"
        )
    else:
        summary_msg += f"📌 Status: {payload.get(ENTRY_STATUS, 'Completed')}\n"
        
    summary_msg += "━━━━━━━━━━━━━━━━━━━━━━━━━━"

    confirm_kb = [
        [InlineKeyboardButton("🚀 Submit Form", callback_data="f_final_submit")],
        [InlineKeyboardButton("❌ Cancel", callback_data="cancel_action")]
    ]
    await target_message.reply_text(summary_msg, reply_markup=InlineKeyboardMarkup(confirm_kb), parse_mode="Markdown")

# ==========================================
# 10. TEXT MESSAGE HANDLER FOR FORMS INPUT
# ==========================================
async def message_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id not in USER_FORM_STATES: return

    state_data = USER_FORM_STATES[chat_id]
    step = state_data.get('step')
    text = update.message.text.strip()

    if step == 'WAITING_FOR_CUSTOM_BANK_NAME':
        state_data['extracted_payload'][ENTRY_BANK] = text
        if state_data['extracted_payload'].get(ENTRY_TYPE2) == 'pm':
            state_data['step'] = 'PM_WAITING_FOR_BRANCH_NAME'
        else:
            state_data['step'] = 'WAITING_FOR_BRANCH_NAME'
        await update.message.reply_text("🏢 *Please enter the Branch Name:*", parse_mode="Markdown")
        return

    if step == 'WAITING_FOR_BRANCH_NAME':
        state_data['extracted_payload'][ENTRY_BRANCH] = text
        state_data['step'] = 'ASK_TYPE'
        
        type_kb = [
            [InlineKeyboardButton("📱 Phone", callback_data="ftype_Phone"),
             InlineKeyboardButton("🏃 Physical", callback_data="ftype_Physical")],
            [InlineKeyboardButton("❌ Cancel Process", callback_data="cancel_action")]
        ]
        await update.message.reply_text("📞 *Select Support Type:*", reply_markup=InlineKeyboardMarkup(type_kb), parse_mode="Markdown")
        return

    if step == 'ASK_ISSUE':
        state_data['extracted_payload'][ENTRY_CASE_ISSUE] = text
        state_data['step'] = 'ASK_STATUS'

        status_kb = [
            [InlineKeyboardButton("✅ Completed", callback_data="fstat_Completed"),
             InlineKeyboardButton("⏳ Pending / On going", callback_data="fstat_On going")],
            [InlineKeyboardButton("❌ Cancel Process", callback_data="cancel_action")]
        ]
        await update.message.reply_text("📌 *Select Case Status:*", reply_markup=InlineKeyboardMarkup(status_kb), parse_mode="Markdown")
        return

    if step == 'WAITING_FOR_PART_NAME':
        state_data['extracted_payload'][ENTRY_PART_NAME] = text
        state_data['step'] = 'PREVIEW_READY'
        await render_summary_and_confirm(update.message, state_data)
        return

    if step == 'PM_WAITING_FOR_BRANCH_NAME':
        state_data['extracted_payload'][ENTRY_BRANCH] = text
        state_data['step'] = 'PREVIEW_READY'
        await render_summary_and_confirm(update.message, state_data)
        return

# ==========================================
# 11. STARTUP MENU INITIALIZER
# ==========================================
async def post_init(application: Application) -> None:
    commands = [
        BotCommand("start", "Initialize bot profile"),
        BotCommand("pending", "View open and unresolved cases"),
        BotCommand("daily", "View daily technician cases report"),
        BotCommand("report", "View weekly technician performance metrics"),
        BotCommand("summary", "View overall weekly summary metrics"),
        BotCommand("export", "Generate incident logs Excel sheet")
    ]
    await application.bot.set_my_commands(commands)
    asyncio.create_task(start_independent_alarm_loop(application.bot))

# ==========================================
# 12. ENGINE INITIATION
# ==========================================
def main():
    if not BOT_TOKEN:
        logger.error("SYSTEM ERROR: TELEGRAM_BOT_TOKEN is missing.")
        return

    threading.Thread(target=run_health_server, daemon=True).start()

    application = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("pending", pending_command))
    application.add_handler(CommandHandler("daily", daily_command))
    application.add_handler(CommandHandler("report", report_command))
    application.add_handler(CommandHandler("summary", summary_command))
    application.add_handler(CommandHandler("export", export_command))
    application.add_handler(CallbackQueryHandler(button_click_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_input_handler))

    application.run_polling()

if __name__ == '__main__':
    main()
